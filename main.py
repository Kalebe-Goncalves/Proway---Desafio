import bd
from openpyxl import Workbook
from openpyxl import load_workbook

"""
Código responsável por realizar o controle da aplicação, realizar ações e mudanças de valores, contém fuções responsáveis pelo funcionamento da lógica do app.
"""


def atualizar_min(placar, min_temp, q_recor_min):
    """
    Função que atualiza o recorde mínimo da temporada.

    placar = int
    min_temp = int
    q_recor_min = int
    """
    lista_min = [
    ]      # Lista base onde serão armazenados localmente (primariamente) os valores do recorde mínimo da temporada
    if placar < min_temp or min_temp == 0:      # Condição para atualizar o valor do recorde mínimo da temporada
        # Condição que verifica se é a primeira entrada (primeira partida armazenada)
        if not min_temp == 0 and placar < min_temp:
            q_recor_min += 1
            lista_min.append(q_recor_min)
        min_temp = placar
        lista_min.append(min_temp)
        return lista_min
    else:
        return lista_min


def atualizar_max(placar, max_temp, q_recor_max):
    """
    Função que atualiza o recorde máximo da temporada. 

    placar = int
    max_temp = int
    q_recor_max = int
    """
    # Mesma lógica da função atualizar_min
    lista_max = []
    if placar > max_temp or max_temp == 0:
        if not max_temp == 0 and placar > max_temp:
            q_recor_max += 1
            lista_max.append(q_recor_max)
        max_temp = placar
        lista_max.append(max_temp)
        return lista_max
    else:
        return lista_max


def gravar_dados_bd(jogo, placar, minimo, maximo, rec_min, rec_max):
    """
    Função responsável por inserir no banco de dados os valores da partida.

    jogo = int
    placar = int
    minimo = int
    maximo = int
    rec_min = int
    rec_max = int
    """
    try:
        # cria uma nova instância no banco de dados e grava os valores
        partida = bd.Partida.create(num_partida=jogo, placar=placar, minimo=minimo,
                                    maximo=maximo, recorde_min=rec_min, recorde_max=rec_max)
        return True
    except (ValueError, TypeError, AssertionError):     # tratamento de erro
        return False


def retornar_valores_bd():
    """
    Função que acessa o banco de dados e retorna os valores contidos nele.
    """
    partidas = bd.Partida.select()
    return partidas     # varaiável tipo <class> ModelSelect


def inserir_dados(placar, jogo, min_temp, max_temp, q_recor_min, q_recor_max):
    """
    Função que contém a lógica para substituir os recordes mínimo e máximo quando quebrados e atualiza quantas vezes estes foram quebrados na temporada.
    A partir do momento em que a lógica é implementa e os valores são atualizados estes são gravados no banco de dados através do método gravar_dados_bd.

    jogo = int
    placar = int
    minimo = int
    maximo = int
    rec_min = int
    rec_max = int
    """
    jogo += 1  # incrementa automaticamente sempre que uma nova partida é inserida pelo usuário
    lista_min = atualizar_min(placar, min_temp, q_recor_min)
    lista_max = atualizar_max(placar, max_temp, q_recor_max)

    if len(lista_min) == 2:     # condição que verifica se há quebra de recorde ou não, tendo quebra o valor minimo da temporada é substituido
        q_recor_min = lista_min[0]
        min_temp = lista_min[1]
    elif len(lista_min) == 1:   # condição que age em conjunto com a anterior, permitindo manter o valor mínimo da partida anterior
        min_temp = lista_min[0]
    if len(lista_max) == 2:     # condição que verifica se há quebra de recorde ou não, tendo quebra o valor MÁXIMO da temporada é substituido
        q_recor_max = lista_max[0]
        max_temp = lista_max[1]
    elif len(lista_max) == 1:   # condição que age em conjunto com a anterior, permitindo manter o valor MÁXIMO da partida anterior
        max_temp = lista_max[0]

    # condição que serve como verificação final para gravar no bd
    if gravar_dados_bd(jogo, placar, min_temp, max_temp, q_recor_min, q_recor_max):
        return True
    else:
        return False


def exportar_planilha_excel():
    """
    Função responsável por acessar, atualizar e salvar a planilha em excel que contém os dados contidos no Banco de dados.
    """
    name_arq = "pontuacao_basqueteMaria.xlsx"   # nome arquivo da planilha já existente
    arquivo_excel = load_workbook(name_arq)     # leitura da planilha
    planilha = arquivo_excel.get_sheet_by_name(
        "Registro_Pontuacao")    # escolha da tabela
    for partida in retornar_valores_bd():   # lógica responsável por colocar os valores em seus respectivos campos (coluna e linha) na tabela
        for coluna in range(1, 7):
            linha = partida.num_partida+1
            if coluna == 1:
                # insere o valor na tabela
                planilha.cell(row=linha, column=coluna,
                              value=partida.num_partida)
            if coluna == 2:
                planilha.cell(row=linha, column=coluna, value=partida.placar)
            if coluna == 3:
                planilha.cell(row=linha, column=coluna, value=partida.minimo)
            if coluna == 4:
                planilha.cell(row=linha, column=coluna, value=partida.maximo)
            if coluna == 5:
                planilha.cell(row=linha, column=coluna,
                              value=partida.recorde_min)
            if coluna == 6:
                planilha.cell(row=linha, column=coluna,
                              value=partida.recorde_max)

    try:
        arquivo_excel.save(name_arq)    # salva as atualizações na planilha
        return True
    except PermissionError:     # tratamento de erro caso a planilha esteja aberta
        return False
